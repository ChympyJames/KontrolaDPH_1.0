import streamlit as st
import pandas as pd
import re
import time
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def get_driver():
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')

    # Locate system-installed ChromeDriver
    chromedriver_path = shutil.which("chromedriver")
    if not chromedriver_path:
        st.error(
            "❌ ChromeDriver not found. Ensure it's installed via `packages.txt`.")
        return None

    service = Service(chromedriver_path)
    driver = webdriver.Chrome(service=service, options=options)
    return driver


# --- Validate Bank Account Format ---
def is_valid_account(account):
    return bool(re.match(r'^[0-9/-]+$', str(account))) if account else False


# --- Split into Batches (2 DIČs per Batch) ---
def split_into_batches(lst, batch_size=1):
    return [lst[i:i + batch_size] for i in range(0, len(lst), batch_size)]


# --- Fetch "Nespolehlivý plátce" Status ---
def fetch_nespolehlivy(driver):
    try:
        # Get all text content from the page
        body_text = driver.find_element(By.TAG_NAME, "body").text

        # Look for the line that contains 'Nespolehlivý plátce:'
        for line in body_text.split('\n'):
            if "Nespolehlivý plátce:" in line:
                # Example line: "Nespolehlivý plátce:    NE"
                parts = line.split(':')
                if len(parts) >= 2:
                    status = parts[1].strip().upper()
                    return [status]

        # If not found in the lines
        return ["NEZNÁMÝ"]

    except Exception as e:
        return ["NEZNÁMÝ"]


# --- Format Excel File ---
def format_excel(output_filename):
    from openpyxl import load_workbook
    wb = load_workbook(output_filename)
    ws = wb.active
    table_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="ResultsTable", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
                                          showFirstColumn=False,
                                          showLastColumn=False,
                                          showRowStripes=True,
                                          showColumnStripes=False)
    ws.add_table(table)
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            len(str(cell.value)) for cell in col) + 2
    wb.save(output_filename)


# --- Process Uploaded File ---
def process_file(uploaded_file):
    df = pd.read_excel(uploaded_file,
                       dtype={"Číslo bank. účtu": str, "Směr.kód": str})

    if df.empty:
        st.error("⚠️ S Excel souborem je něco špatně.")
        return None

    # Apply filtering (PREVOD + DIČ starts with "CZ")
    df = df[(df["Forma úhrady"] == "PREVOD") & (
        df["DIČ"].astype(str).str.startswith("CZ", na=False))]

    df = df[
        (df['Stav úhrady dokladu'].isnull())]

    # Format Bank Account
    df["Směr.kód"] = df["Směr.kód"].str.extract(r"(\d+)")[0].fillna(
        "0000").astype(str).str.zfill(4)
    df["Číslo bank. účtu"] = df["Číslo bank. účtu"].astype(str).str.extract(r"([\d\-]+)")[0]
    df["Bankovní účet"] = df["Číslo bank. účtu"] + "/" + df["Směr.kód"]

    # Initialize Output File
    output_filename = f"Kontrola_ucty_DPH_{datetime.now().strftime('%d-%m-%Y_%H%M')}.xlsx"
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(
        ["DIČ", "Bankovní účet", "Název firmy nebo jméno osoby", "Match",
         "Nespolehlivý plátce"])

    driver = get_driver()
    driver.get("https://adisspr.mfcr.cz/dpr/DphReg")

    dic_batches = split_into_batches(df['DIČ'].tolist(), 1)
    total_batches = len(dic_batches)

    # Progress Bar + Status Message
    progress_bar = st.progress(0)
    status_text = st.empty()
    time_text = st.empty()

    start_time = time.time()  # Start timing

    for batch_idx, batch in enumerate(dic_batches):
        batch_size = len(batch)
        current_dic = ", ".join(batch)

        driver.get("https://adisspr.mfcr.cz/dpr/DphReg")

        # Enter DIČs
        input_fields = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "input[id^='form:dt']")))
        for i, dic_number in enumerate(batch):
            if i < len(input_fields):
                input_fields[i].send_keys(dic_number.replace("CZ", ""))

        # Click Search
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "form:hledej"))).click()

        # Scrape Bank Accounts
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "table[id^='tableUcty'] tbody")))
            tables = driver.find_elements(By.CSS_SELECTOR,
                                          "table[id^='tableUcty'] tbody")
            scraped_accounts = [line.split()[0] for table in tables for line in
                                table.text.splitlines()]
        except (NoSuchElementException, TimeoutException):
            scraped_accounts = None

        # Fetch "Nespolehlivý plátce"
        nespolehlivy_list = fetch_nespolehlivy(driver)

        # Calculate estimated time left
        elapsed_time = time.time() - start_time
        avg_time_per_batch = elapsed_time / (
                    batch_idx + 1) if batch_idx > 0 else 0
        remaining_batches = total_batches - (batch_idx + 1)
        estimated_time_left = avg_time_per_batch * remaining_batches

        percentage_done = int(((batch_idx + 1) / total_batches) * 100)

        # Update UI
        status_text.text(
            f"🔍 Zpracovávám várku: {batch_idx + 1}/{total_batches} | DIČ: {current_dic}")
        time_text.text(f"⏳ Zbývá: {estimated_time_left:.2f} sec")
        progress_bar.progress(percentage_done / 100)

        # Match & Save Results
        for i, dic_number in enumerate(batch):
            row = df[df['DIČ'] == dic_number].iloc[0]
            bank_account = str(row["Bankovní účet"])
            company_name = str(row["Název firmy nebo jméno osoby"])

            if scraped_accounts is None:
                account_check_result = "Nenalezen účet"

            elif not is_valid_account(bank_account):
                account_check_result = "Chyba zadání"

            elif bank_account in scraped_accounts:
                account_check_result = "✔"

            else:
                account_check_result = "Neshoda účtu"

            new_ws.append(
                [dic_number, bank_account, company_name, account_check_result,
                 nespolehlivy_list[i]])

    driver.quit()
    new_wb.save(output_filename)
    format_excel(output_filename)
    return output_filename

def reset_app():
    try:
        driver.quit()
    except Exception:
        pass
    st.cache_data.clear()
    st.session_state.reset_after_download = False
    st.experimental_rerun()

def main():
    st.set_page_config(page_title="🔍 Kontrol účtů pro účely DPH 🔍",
                       page_icon="✅", layout="centered")
    st.title("🔍 Kontrol účtů pro účely DPH 🔍")

    uploaded_file = st.file_uploader("📂 Nahrajte Excel soubor", type=["xlsx"])
    if uploaded_file and st.button("🔍 Spustit kontrolu"):
        with st.spinner("⏳ Zpracovávám data..."):
            output_filename = process_file(uploaded_file)

        if output_filename:
            st.success(
                f"✅ Kontrola dokončena! Výsledky uloženy jako: {output_filename}")

            # Open the file and provide a download button
            with open(output_filename, "rb") as file:
                if st.download_button(
                    label="📥 Stáhnout výsledky",
                    data=file,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ):
                    st.success("✅ Aplikace se resetuje...")
                    driver.quit()  # Optional safety if not already quit, if driver is global
                    reset_app()   # Soft reboot the app


# --- Run Main Function ---
if __name__ == "__main__":
    main()
